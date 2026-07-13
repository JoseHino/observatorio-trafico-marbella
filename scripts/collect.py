# -*- coding: utf-8 -*-
"""
Colector de datos municipales de tráfico de la DGT para Marbella.

Descarga los ficheros oficiales "DGT en cifras - Información municipal"
(un Excel por año con TODOS los municipios de España), extrae Marbella y
calcula agregados comparativos (Málaga capital, provincia de Málaga,
Andalucía y España). Genera:

  - data/trafico_marbella.json   -> alimenta el dashboard
  - Observatorio_Trafico_Marbella_DGT.xlsx -> Excel enriquecido

Fuente: https://www.dgt.es/menusecundario/dgt-en-cifras/
        (Tema: Información municipal)

No requiere API key ni login: son descargas HTTP directas y estables.
"""
import hashlib
import io
import json
import re
import sys
import time
import unicodedata
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen
from urllib.error import HTTPError

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
DOCS = ROOT / "docs"
DATA.mkdir(exist_ok=True)
DOCS.mkdir(exist_ok=True)

BASE = ("https://www.dgt.es/export/sites/web-DGT/.galleries/downloads/"
        "dgt-en-cifras/informacion_municipal")

# dataset -> (carpeta, prefijo de fichero, rango de años a intentar)
DATASETS = {
    "general": ("Datos-municipales-general", "DatosMunicipalesGeneral"),
    "siniestralidad": ("Datos-municipales-siniestralidad", "DatosMunicipalesSiniestralidad"),
    "sanciones": ("Datos-municipales-sanciones-y-puntos", "DatosMunicipalesSancionesPuntos"),
}
YEARS = list(range(2015, datetime.now(timezone.utc).year + 1))

MUNICIPIO = "Marbella"
PROVINCIA = "Málaga"
COMUNIDAD = "Andalucía"
UA = "Mozilla/5.0 (compatible; ObservatorioMarbella/1.0)"


def strip_accents(s):
    if s is None:
        return ""
    s = str(s)
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c)).strip().lower()


def num(x):
    if x is None or x == "":
        return None
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def fetch(url, retries=3):
    for attempt in range(retries):
        try:
            req = Request(url, headers={"User-Agent": UA})
            with urlopen(req, timeout=60) as resp:
                return resp.read()
        except HTTPError as e:
            if e.code == 404:
                return None
            if attempt == retries - 1:
                raise
            time.sleep(2)
        except Exception:
            if attempt == retries - 1:
                raise
            time.sleep(2)
    return None


_MINIMAL_CORE = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
    b'xmlns:dc="http://purl.org/dc/elements/1.1/" '
    b'xmlns:dcterms="http://purl.org/dc/terms/" '
    b'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">'
    b'<dcterms:created xsi:type="dcterms:W3CDTF">2020-01-01T00:00:00Z</dcterms:created>'
    b'<dcterms:modified xsi:type="dcterms:W3CDTF">2020-01-01T00:00:00Z</dcterms:modified>'
    b'</cp:coreProperties>')


def sanitize_xlsx(raw):
    """Reescribe docProps/core.xml con fechas válidas (algunos ficheros DGT
    traen fechas mal formateadas que rompen openpyxl)."""
    src = zipfile.ZipFile(io.BytesIO(raw))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "docProps/core.xml":
                data = _MINIMAL_CORE
            dst.writestr(item, data)
    return buf.getvalue()


def load_sheet(raw):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except ValueError:
        wb = openpyxl.load_workbook(io.BytesIO(sanitize_xlsx(raw)),
                                    data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # localizar fila de cabecera (contiene "Municipio")
    hdr_idx = 0
    for i, r in enumerate(rows[:6]):
        if r and any(strip_accents(c) == "municipio" for c in r if c is not None):
            hdr_idx = i
            break
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_idx]]
    body = rows[hdr_idx + 1:]
    return header, body


def numeric_columns(header):
    """Columnas de datos = todas menos las de identificación."""
    skip = {"codigo ine", "codigo", "municipio", "provincia",
            "comunidad autonoma", "ano", "año"}
    return [i for i, h in enumerate(header) if strip_accents(h) not in skip and h]


def aggregate(header, body):
    ncols = numeric_columns(header)
    iMuni = next((i for i, h in enumerate(header) if strip_accents(h) == "municipio"), 1)
    iProv = next((i for i, h in enumerate(header) if strip_accents(h) == "provincia"), 2)
    iCa = next((i for i, h in enumerate(header) if strip_accents(h) == "comunidad autonoma"), 3)

    scopes = {
        "Marbella": lambda r: strip_accents(r[iMuni]) == strip_accents(MUNICIPIO)
        and strip_accents(r[iProv]) == strip_accents(PROVINCIA),
        "Málaga capital": lambda r: strip_accents(r[iMuni]) == strip_accents(PROVINCIA)
        and strip_accents(r[iProv]) == strip_accents(PROVINCIA),
        "Provincia Málaga": lambda r: strip_accents(r[iProv]) == strip_accents(PROVINCIA),
        "Andalucía": lambda r: strip_accents(r[iCa]) == strip_accents(COMUNIDAD),
        "España": lambda r: True,
    }
    # "Marbella" y "Málaga capital" son filas únicas -> tomamos valores tal cual.
    # El resto se suman (agregados de área).
    singletons = {"Marbella", "Málaga capital"}

    out = {}
    for scope, pred in scopes.items():
        vals = {header[i]: (None if scope in singletons else 0.0) for i in ncols}
        found = False
        for r in body:
            if not r or r[iMuni] is None:
                continue
            if pred(r):
                found = True
                for i in ncols:
                    v = num(r[i])
                    if scope in singletons:
                        vals[header[i]] = v
                        continue
                    if v is not None:
                        vals[header[i]] += v
                if scope in singletons:
                    break
        out[scope] = vals if found else None
    return [header[i] for i in ncols], out


# Balances trimestrales oficiales (4º trimestre = año completo). Cada balance
# publica el año de referencia + el anterior ya consolidado. Se añaden años a
# medida que el Ministerio los publica. El pipeline es tolerante a fallos: si un
# PDF no se puede descargar, conserva lo ya guardado en data/criminalidad.json.
CRIM_BALANCES = {
    2022: "https://www.interior.gob.es/opencms/export/sites/default/.galleries/galeria-de-prensa/documentos-y-multimedia/balances-e-informes/2022/Balance-de-Criminalidad-Cuarto-Trimestre-2022.pdf",
    2024: "https://www.interior.gob.es/opencms/export/sites/default/.galleries/galeria-de-prensa/documentos-y-multimedia/balances-e-informes/2024/BALANCE-CRIMINALIDAD-CUARTO-TRIMESTRE-2024.pdf",
    2025: "https://www.interior.gob.es/opencms/export/sites/default/.galleries/galeria-de-prensa/documentos-y-multimedia/balances-e-informes/2025/Balance-de-Criminalidad_Cuarto_Trimestre_2025.pdf",
}
CRIM_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120 Safari/537.36",
    "Accept": "application/pdf,*/*", "Accept-Language": "es-ES,es;q=0.9",
    "Referer": "https://www.interior.gob.es/",
}


def _crim_clean(label):
    if label.startswith("I.") and "CONVENCIONAL" in label:
        return "Criminalidad convencional (total)"
    if label.startswith("II") and "CIBERCRIMINALIDAD" in label:
        return "Cibercriminalidad (total)"
    if "TOTAL CRIMINALIDAD" in label:
        return "TOTAL criminalidad"
    return re.sub(r"^\s*\d+(\.\d+)?[.\-]+\s*", "", label).strip()


def parse_balance_marbella(raw):
    """Extrae la tabla 'Municipio de Marbella' de un balance de criminalidad
    (PDF). Devuelve {anio: {tipologia: valor}} para los dos años del balance."""
    import fitz
    doc = fitz.open(stream=raw, filetype="pdf")
    is_num = lambda s: bool(re.fullmatch(r"-?\d{1,3}(\.\d{3})*|-?\d+", s))
    for i in range(doc.page_count):
        t = doc[i].get_text()
        if "Municipio de Marbella" not in t or "TIPOLOG" not in t:
            continue
        lines = [l.strip() for l in t.split("\n") if l.strip()]
        hi = next(k for k, l in enumerate(lines) if l.startswith("TIPOLOG"))
        yrs = [l for l in lines[hi + 1:hi + 6] if re.fullmatch(r"20\d\d", l)][:2]
        if len(yrs) < 2:
            continue
        yp, yc = yrs
        data = {yp: {}, yc: {}}
        label, nums = None, []
        for l in lines[hi + 3:]:
            if l.startswith(("Columna", "Página", "INFRACCIONES")):
                break
            if re.fullmatch(r"-?\d+,\d+", l):
                continue  # porcentaje
            if is_num(l):
                if label is not None:
                    nums.append(l)
            else:
                label, nums = l, []
            if label is not None and len(nums) == 2:
                c = _crim_clean(label)
                data[yp][c] = int(nums[0].replace(".", ""))
                data[yc][c] = int(nums[1].replace(".", ""))
                label = None
        doc.close()
        return data
    doc.close()
    return {}


def collect_criminalidad():
    """Criminalidad de Marbella (Ministerio del Interior). Parte del histórico
    ya guardado y lo amplía/consolida con los balances oficiales disponibles.
    Nunca pierde años previos aunque falle la descarga."""
    path = DATA / "criminalidad.json"
    base = json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}
    series = base.get("series", {}).get("Marbella", {})
    colorder = list(base.get("columnas", []))
    try:
        import fitz  # noqa: F401
        for ref in sorted(CRIM_BALANCES):
            url = CRIM_BALANCES[ref]
            try:
                raw = urlopen(Request(url, headers=CRIM_HEADERS), timeout=90).read()
                d = parse_balance_marbella(raw)
                for y in sorted(d):
                    series.setdefault(y, {})
                    for c, v in d[y].items():
                        if c not in colorder:
                            colorder.append(c)
                        series[y][c] = v  # el balance más reciente consolida
                print(f"[criminalidad] balance {ref}: OK ({', '.join(sorted(d))})")
            except Exception as e:
                print(f"[criminalidad] balance {ref}: no disponible ({getattr(e,'code',e)})")
            time.sleep(6)  # petición educada (evita el límite de ráfaga del MIR)
    except ImportError:
        print("[criminalidad] PyMuPDF no instalado; se conserva lo guardado.")
    if not series:
        print("[criminalidad] sin datos.")
        return None
    anios = sorted(int(y) for y in series)
    out = {
        "columnas": colorder,
        "anios": anios,
        "series": {"Marbella": series},
        "fuente": "Ministerio del Interior — Balance trimestral de criminalidad (SEC)",
        "fuente_url": "https://estadisticasdecriminalidad.ses.mir.es/",
        "nota": "Acumulado enero-diciembre. Cada año se consolida en el balance del año siguiente.",
    }
    path.write_text(json.dumps(out, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"[criminalidad] años: {anios} | tipologías: {len(colorder)}")
    return out


def collect_aforos():
    """Aforos de tráfico de las estaciones MA-341 y MA-417 (hoja de seguimiento
    propia, data/aforos.csv). Datos mensuales -> se agregan a totales anuales
    para los años completos, y se conserva el detalle mensual (estacionalidad)."""
    import csv as _csv
    from collections import defaultdict, OrderedDict
    path = DATA / "aforos.csv"
    if not path.exists():
        print("\n[aforos] data/aforos.csv no encontrado; se omite.")
        return None
    COLS = ["MA-341 · Ligeros", "MA-341 · Pesados", "MA-341 · Total",
            "MA-417 · Ligeros", "MA-417 · Pesados", "MA-417 · Total"]
    rows = list(_csv.reader(path.open(encoding="utf-8")))
    body = rows[1:]
    years = defaultdict(lambda: {c: [] for c in COLS})
    periodos = []
    mensual = OrderedDict((c, []) for c in COLS)
    for r in body:
        per = r[0]
        periodos.append(per)
        vals = [int(x) if x not in ("", "None") else None for x in r[1:7]]
        for c, v in zip(COLS, vals):
            mensual[c].append(v)
            years[per[:4]][c].append(v)
    series, anios = {}, []
    for y in sorted(years):
        cols = years[y]
        if not all(len([v for v in cols[c] if v is not None]) == 12 for c in COLS):
            continue  # solo años con los 12 meses
        anios.append(int(y))
        series[y] = {c: sum(v for v in cols[c] if v is not None) for c in COLS}
    print(f"\n[aforos] años completos: {anios} ({len(periodos)} meses en total)")
    return {
        "columnas": COLS,
        "anios": anios,
        "series": {"Marbella": series},
        "mensual": {"periodos": periodos, "valores": mensual},
        "fuente": "Estaciones de aforo MA-341 y MA-417 (seguimiento propio)",
    }


def main():
    result = {
        "meta": {
            "generado": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "municipio": MUNICIPIO,
            "fuente": "DGT - DGT en cifras (Información municipal)",
            "fuente_url": "https://www.dgt.es/menusecundario/dgt-en-cifras/",
            "ambitos": ["Marbella", "Málaga capital", "Provincia Málaga",
                        "Andalucía", "España"],
        },
        "datasets": {},
    }

    for name, (folder, prefix) in DATASETS.items():
        print(f"\n=== {name.upper()} ===")
        columns = None
        years_ok = []
        series = {a: {} for a in result["meta"]["ambitos"]}
        for y in YEARS:
            url = f"{BASE}/{folder}/{prefix}_{y}.xlsx"
            raw = fetch(url)
            if raw is None:
                print(f"  {y}: no publicado (404)")
                continue
            header, body = load_sheet(raw)
            cols, agg = aggregate(header, body)
            columns = cols
            years_ok.append(y)
            for scope, vals in agg.items():
                if vals is not None:
                    series[scope][str(y)] = {c: vals.get(c) for c in cols}
            print(f"  {y}: OK ({len(body)} municipios)")
        result["datasets"][name] = {
            "columnas": columns or [],
            "anios": years_ok,
            "series": series,
        }

    af = collect_aforos()
    if af:
        result["datasets"]["aforos"] = af

    crim = collect_criminalidad()
    if crim:
        result["datasets"]["criminalidad"] = crim

    (DATA / "trafico_marbella.json").write_text(
        json.dumps(result, ensure_ascii=False, indent=1), encoding="utf-8")
    # copia servida por el dashboard
    (DOCS / "trafico_marbella.json").write_text(
        json.dumps(result, ensure_ascii=False), encoding="utf-8")

    # Firma de contenido SOLO de los datos (excluye meta.generado, que cambia
    # en cada ejecución). Sirve para distinguir "datos nuevos reales" del
    # latido mensual y disparar el aviso por email únicamente cuando toca.
    firma = hashlib.sha256(
        json.dumps(result["datasets"], sort_keys=True, ensure_ascii=False).encode("utf-8")
    ).hexdigest()
    (DATA / "datahash.txt").write_text(firma, encoding="utf-8")

    print(f"\n[OK] JSON escrito. Años general={result['datasets']['general']['anios']}")
    print(f"[OK] Firma de datos: {firma[:12]}...")
    return result


if __name__ == "__main__":
    main()
