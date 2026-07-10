# -*- coding: utf-8 -*-
"""
Genera el Excel enriquecido del Observatorio de Tráfico de Marbella a partir
de data/trafico_marbella.json (producido por collect.py).

Hojas:
  - Portada           : metadatos, fuente y leyenda
  - General_Marbella  : serie histórica completa de Marbella
  - Siniestralidad_Marbella
  - Sanciones_Marbella
  - Comparativa_<año> : Marbella vs Málaga capital / provincia / Andalucía / España
  - Indicadores       : tasas derivadas (motorización, parque limpio, siniestralidad...)
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = ROOT / "Observatorio_Trafico_Marbella_DGT.xlsx"

AZUL = "1F4E79"
AZUL_CLARO = "D6E4F0"
GRIS = "808080"
AMBITOS = ["Marbella", "Málaga capital", "Provincia Málaga", "Andalucía", "España"]

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=AZUL)
sub_fill = PatternFill("solid", fgColor=AZUL_CLARO)
title_font = Font(bold=True, size=14, color=AZUL)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def gv(dd, *keys):
    """Devuelve el valor de la columna que casa con alguna clave.
    Prioriza coincidencia exacta (ignorando mayúsculas) sobre subcadena."""
    if not dd:
        return None
    low = {col.lower(): v for col, v in dd.items()}
    for k in keys:                       # 1) exacta
        if k.lower() in low:
            return low[k.lower()]
    for k in keys:                       # 2) subcadena
        for col, v in dd.items():
            if k.lower() in col.lower():
                return v
    return None


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_serie(wb, title, ds, scope="Marbella"):
    ws = wb.create_sheet(title)
    cols = ds["columnas"]
    years = ds["anios"]
    series = ds["series"].get(scope, {})
    ws.cell(1, 1, title.replace("_", " ")).font = title_font
    ws.cell(2, 1, f"Ámbito: {scope}  ·  Fuente: DGT en cifras (Información municipal)").font = Font(italic=True, color=GRIS)
    hrow = 4
    ws.cell(hrow, 1, "Indicador").value = "Indicador"
    for j, y in enumerate(years, start=2):
        ws.cell(hrow, j, y)
    style_header(ws, hrow, len(years) + 1)
    for i, col in enumerate(cols, start=hrow + 1):
        ws.cell(i, 1, col).font = Font(bold=True)
        ws.cell(i, 1).fill = sub_fill
        ws.cell(i, 1).border = border
        for j, y in enumerate(years, start=2):
            v = (series.get(str(y)) or {}).get(col)
            cell = ws.cell(i, j, v)
            cell.border = border
            if isinstance(v, float) and v.is_integer():
                cell.value = int(v)
            cell.number_format = "#,##0.###"
    ws.freeze_panes = "B5"
    autosize(ws, [42] + [11] * len(years))
    return ws


def sheet_comparativa(wb, data, year):
    ws = wb.create_sheet(f"Comparativa_{year}")
    ws.cell(1, 1, f"Comparativa {year} — Marbella vs entorno").font = title_font
    ws.cell(2, 1, "Cifras absolutas por ámbito (misma fuente DGT).").font = Font(italic=True, color=GRIS)
    g = data["datasets"]["general"]["series"]
    s = data["datasets"]["siniestralidad"]["series"]
    row = 4
    kpis = [
        ("Población total", g, ("Población Total", "Poblaci")),
        ("Censo de conductores", g, ("Censo Conductores",)),
        ("Parque total de vehículos", g, ("Parque Total",)),
        ("  · Turismos", g, ("Parque Turismos",)),
        ("  · Motocicletas", g, ("Parque Motocicletas",)),
        ("Vehículos electrificados", g, ("Electrificado",)),
        ("Distintivo ambiental 0", g, ("Distintivo 0",)),
        ("Distintivo ambiental ECO", g, ("Distintivo ECO",)),
        ("Sin distintivo ambiental", g, ("Sin Distintivo",)),
        ("Accidentes con víctimas", s, ("Nº Accidentes con Víctimas", "Accidentes con V")),
        ("Fallecidos", s, ("Fallecidos",)),
        ("Peatones heridos (leves)", s, ("Peatones Heridos Leves",)),
    ]
    ws.cell(row, 1, "Indicador")
    for j, a in enumerate(AMBITOS, start=2):
        ws.cell(row, j, a)
    style_header(ws, row, len(AMBITOS) + 1)
    r = row + 1
    for label, src, keys in kpis:
        ws.cell(r, 1, label).font = Font(bold=label.startswith(("P", "C", "V", "A", "F")))
        ws.cell(r, 1).border = border
        for j, a in enumerate(AMBITOS, start=2):
            v = gv(src.get(a, {}).get(str(year), {}), *keys)
            cell = ws.cell(r, j, int(v) if isinstance(v, float) and v.is_integer() else v)
            cell.number_format = "#,##0"
            cell.border = border
        r += 1
    autosize(ws, [30] + [16] * len(AMBITOS))
    ws.freeze_panes = "B5"


def sheet_indicadores(wb, data):
    ws = wb.create_sheet("Indicadores")
    ws.cell(1, 1, "Indicadores derivados de Marbella").font = title_font
    ws.cell(2, 1, "Tasas calculadas sobre los datos DGT (comparables en el tiempo).").font = Font(italic=True, color=GRIS)
    g = data["datasets"]["general"]
    s = data["datasets"]["siniestralidad"]
    years = g["anios"]
    gm = g["series"]["Marbella"]
    sm = s["series"]["Marbella"]

    def row_calc(label, fn):
        return (label, fn)

    defs = [
        ("Vehículos por 1.000 habitantes",
         lambda y: safe(gv(gm.get(y, {}), "Parque Total"), gv(gm.get(y, {}), "Poblaci"), 1000)),
        ("Turismos por 1.000 habitantes",
         lambda y: safe(gv(gm.get(y, {}), "Parque Turismos"), gv(gm.get(y, {}), "Poblaci"), 1000)),
        ("% conductores sobre población",
         lambda y: safe(gv(gm.get(y, {}), "Censo Conductores"), gv(gm.get(y, {}), "Poblaci"), 100)),
        ("% parque ECO+0 (limpio)",
         lambda y: safe(add(gv(gm.get(y, {}), "Distintivo ECO"), gv(gm.get(y, {}), "Distintivo 0")),
                        gv(gm.get(y, {}), "Parque Total"), 100)),
        ("% parque sin distintivo",
         lambda y: safe(gv(gm.get(y, {}), "Sin Distintivo"), gv(gm.get(y, {}), "Parque Total"), 100)),
        ("Antigüedad media del parque (años)",
         lambda y: gv(gm.get(y, {}), "Antig", "Antigüedad Media del Parque")),
        ("Accidentes con víctimas por 10.000 hab.",
         lambda y: safe(gv(sm.get(y, {}), "Nº Accidentes con Víctimas", "Accidentes con V"), gv(gm.get(y, {}), "Poblaci"), 10000)),
        ("Fallecidos por 100.000 hab.",
         lambda y: safe(gv(sm.get(y, {}), "Fallecidos"), gv(gm.get(y, {}), "Poblaci"), 100000)),
    ]
    hrow = 4
    ws.cell(hrow, 1, "Indicador")
    for j, y in enumerate(years, start=2):
        ws.cell(hrow, j, y)
    style_header(ws, hrow, len(years) + 1)
    for i, (label, fn) in enumerate(defs, start=hrow + 1):
        ws.cell(i, 1, label).font = Font(bold=True)
        ws.cell(i, 1).fill = sub_fill
        ws.cell(i, 1).border = border
        for j, y in enumerate(years, start=2):
            v = fn(str(y))
            cell = ws.cell(i, j, round(v, 2) if isinstance(v, (int, float)) else v)
            cell.number_format = "#,##0.00"
            cell.border = border
    autosize(ws, [40] + [10] * len(years))
    ws.freeze_panes = "B5"


def safe(a, b, k=1):
    if a is None or b in (None, 0):
        return None
    return a / b * k


def add(a, b):
    if a is None and b is None:
        return None
    return (a or 0) + (b or 0)


def sheet_portada(wb, data):
    ws = wb.create_sheet("Portada", 0)
    ws.sheet_view.showGridLines = False
    ws.cell(2, 2, "OBSERVATORIO DE TRÁFICO DE MARBELLA").font = Font(bold=True, size=18, color=AZUL)
    ws.cell(3, 2, "Datos municipales DGT · Serie histórica y comparativas").font = Font(size=12, color=GRIS)
    meta = data["meta"]
    info = [
        ("Municipio", meta["municipio"]),
        ("Fuente", meta["fuente"]),
        ("URL fuente", meta["fuente_url"]),
        ("Generado", meta["generado"]),
        ("Ámbitos de comparación", ", ".join(meta["ambitos"])),
        ("", ""),
        ("Dataset", "Años disponibles"),
        ("Información general (parque, censo, distintivos)",
         rng(data["datasets"]["general"]["anios"])),
        ("Siniestralidad (accidentes, fallecidos, heridos)",
         rng(data["datasets"]["siniestralidad"]["anios"])),
        ("Sanciones y puntos", rng(data["datasets"]["sanciones"]["anios"])),
    ]
    r = 5
    for k, v in info:
        ws.cell(r, 2, k).font = Font(bold=True, color=AZUL) if k else Font()
        ws.cell(r, 3, v)
        r += 1
    autosize(ws, [3, 48, 55])


def rng(years):
    return f"{min(years)}–{max(years)} ({len(years)} años)" if years else "—"


def main():
    data = json.loads((DATA / "trafico_marbella.json").read_text(encoding="utf-8"))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_portada(wb, data)
    sheet_serie(wb, "General_Marbella", data["datasets"]["general"])
    sheet_serie(wb, "Siniestralidad_Marbella", data["datasets"]["siniestralidad"])
    sheet_serie(wb, "Sanciones_Marbella", data["datasets"]["sanciones"])
    last_general = max(data["datasets"]["general"]["anios"])
    sheet_comparativa(wb, data, last_general)
    sheet_indicadores(wb, data)
    wb.save(OUT)
    # copia descargable desde el sitio (GitHub Pages sirve /docs)
    wb.save(ROOT / "docs" / OUT.name)
    print(f"[OK] Excel escrito: {OUT}")


if __name__ == "__main__":
    main()
